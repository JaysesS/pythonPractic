import subprocess
import string
import re
import openpyxl
from datetime import datetime
import os
import sys
import xlsxwriter
import getpass
import paramiko
import socket

class Grab():

    def __init__(self, progress_signal, users_list, path_to_save_report):

        self.USER = 'farm'
        self.SECRET = getpass.getpass("Password for %s: \n" % USER)
        self.port = 22

        self.progress_signal = progress_signal
        self.ssh = paramiko.SSHClient()
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        self.host_at_moment = ''

        self.path_to_save_report = path_to_save_report
        self.hosts_list = users_list
        
    def get_output_ssh(self, command):
        self.ssh.connect(hostname=self.host_at_moment, username=self.USER, password=self.SECRET, port=self.port, timeout=10.0, auth_timeout= 10.0)
        printable = set(string.printable)
        try:
            stdin, stdout, stderr = self.ssh.exec_command(command)
            out = stdout.read().decode()
            self.ssh.close()
            return ''.join(list(filter(lambda x: x in printable, out)))
        except Exception:
            print('Some problem with command')
            return -1

    def get_motherboard(self):
        out = self.get_output_ssh('inxi -M').replace('12', '')
        out = out.split('Mobo')[1].split('Bios')[0].lstrip(' ')
        return out

    def parse_out_inxi(self, out):
        mother = self.get_motherboard()
        out = out.strip().split('~')
        info = dict()

        if len(out) == 1:
            out = out[0].strip().split(':')
            fx = self.get_output_ssh('hostname').strip()
            cpu = out[1].split('\n')[0].replace('speed/min/max', '').strip().replace('(-MT MCP MCM-)', '')
            speed = out[2].split('/')[2].replace('Kernel', '').strip()
            kernel = out[3].split('[')[0].replace('Up', '').strip()
            mother = mother.strip().split('serial')[0].replace('model ', '')
            mem = out[5].split('(')[0].split('/')[1].strip()
            descript = ''
        else:
            try:
                fx = self.get_output_ssh('hostname').strip()
                cpu = out[1].split('[')[0].replace('speed/max','').replace('(-HT-MCP-)', '')
                speed = out[2].split('[')[0].split('/')[1]
                kernel = out[3].split('[')[0].replace('Up', '')
                mother = mother.strip().replace('model ', '')
                mem = out[5].split('[')[0].split('/')[1].replace('HDD', '')
                descript = ''
            except IndexError:
                fx = self.get_output_ssh('hostname').strip()
                cpu = out[1].split('(')[0]
                speed = out[2].split('(')[0]
                kernel = out[3].split('[')[0].replace(' Up', '')
                mother = mother.strip().replace('model ', '')
                mem = out[5].split('/')[1].split(' ')[0]
                descript = ''

        info = {
                'fx': fx,
                'cpu': cpu,
                'speed': speed,
                'kernel': kernel,
                'mother': mother,
                'mem': mem,
                'descript': descript
            }
        return info

    def parse_out_nvidia(self, out):
        out = out.split('  ')
        out = [x.replace('\n', '')
                .replace('+', '')
                .replace('-', '')
                .replace(' ', '')
                .replace('|', '')
                .replace('=', '') 
                for x in out if x]
        
        info = {
            'driver_version' : out[2].replace("DriverVersion:", ''),
            'gpu': out[13],
            'size' : out[22] #25
        }

        if info['gpu'] == 'ComputeM.':
            info['gpu'] = out[15]
            info['size'] = out[25]

        if info['gpu'] == '0':
            info['gpu'] = out[14]
            info['size'] = out[25]

        if '%' in info['size']:
            info['size'] = out[24]
            
        if 'GeForceGTX106' in info['gpu']:
            info['gpu'] = 'GeForceGTX1060'

        if 'GeForceRTX207' in info['gpu']:
            info['gpu'] = 'GeForceRTX2070'
        
        info['size'] = info['size'].replace('/', '')
        
        return info

    def check_version_linux(self):
        try:
            version = self.get_output_ssh('cat /etc/issue').strip().replace('\\n', '').replace('\\l', '')
            if '19' in version:
                return 19
            elif '18' in version:
                return 18
        except Exception:
            print('\tIt was not possible to determine the Linux version or the version is not equal to 18, 19.')
            return -1

    def modify_list_for_excel(self, system, video):
        a = list()

        a.append('user')
        a.append(system['fx'])
        a.append(system['cpu'])
        if video != 0:
            a.append(video['gpu'] + ' ' + video['size'])
        else:
            a.append(' ')
        a.append(system['mother'])
        a.append(system['mem'])
        a.append(system['kernel'])
        a.append('ip')
        a.append(system['descript'])

        return a

    def styled_append_header(self, data, page):
        double = openpyxl.styles.Side(border_style="double", color="000000")
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        for c in data:
            c = openpyxl.cell.cell.Cell(page, column="A", row= 1, value=c)
            c.font = openpyxl.styles.Font(size=10,bold=True)
            c.alignment = openpyxl.styles.Alignment(horizontal='center')
            c.fill = openpyxl.styles.PatternFill("solid", fgColor='00faff')
            c.border = openpyxl.styles.Border(top=double, bottom=double, left=thin, right=thin)  
            yield c

    def styled_append_data_error(self, data, page):
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        for c in data:
            c = openpyxl.cell.cell.Cell(page, column="A", row= 1, value=c)
            c.font = openpyxl.styles.Font(size=10,bold=True)
            c.fill = openpyxl.styles.PatternFill("solid", fgColor='fe8989')
            c.alignment = openpyxl.styles.Alignment(horizontal='center')
            c.border = openpyxl.styles.Border(top=thin, bottom=thin, right=thin, left=thin)
            yield c

    def styled_append_data(self, data, page):
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        for c in data:
            c = openpyxl.cell.cell.Cell(page, column="A", row= 1, value=c)
            c.font = openpyxl.styles.Font(size=10,bold=True)
            c.fill = openpyxl.styles.PatternFill("solid", fgColor='eeff00')
            c.alignment = openpyxl.styles.Alignment(horizontal='center')
            c.border = openpyxl.styles.Border(top=thin, bottom=thin, right=thin, left=thin)
            yield c

    def write_data_excel(self, wb, page, system, video, col = None):
        if system != 0 and video != 0:
            page.append(self.styled_append_data(self.modify_list_for_excel(system, video), page))
        elif system != 0 and video == 0:
            a = [' '] * col
            a[0] = 'user'
            a[1] = self.host_at_moment
            if "ERROR" in system:
                a[col-1] = system
                page.append(self.styled_append_data_error(a , page))
            elif len(system) == 0:
                print('Dependency ERROR')
                a[col-1] = 'CHECK INXI ON PC'
                page.append(self.styled_append_data_error(a , page))
            else:
                page.append(self.styled_append_data(self.modify_list_for_excel(system, 0), page))
        
        wb.save(self.path_to_save_report)

    def write_main_and_fx_header(self):

        wb = openpyxl.load_workbook(filename = self.path_to_save_report)
        page = wb.active

        header_main = [ 'USER', 'HOST', 'CPU', 'GPU','MOTHERBOARD', 'RAM', 'KERNEL' ,'IP' , 'DESCRIPTION']
        page.append(self.styled_append_header(header_main, page))

        header_fx = page['A2']
        header_fx.value = 'FX'
        header_fx.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        header_fx.font = openpyxl.styles.Font(size=12, bold=True)
        header_fx.fill = openpyxl.styles.PatternFill('solid', fgColor="e9a5f2")
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        header_fx.border = openpyxl.styles.Border(left=thin, right=thin)

        page.merge_cells('A2:I2')

        page.column_dimensions["A"].width = 15.0
        page.column_dimensions["B"].width = 10.0
        page.column_dimensions["C"].width = 45.0
        page.column_dimensions["D"].width = 30.0
        page.column_dimensions["E"].width = 45.0
        page.column_dimensions["F"].width = 20.0
        page.column_dimensions["G"].width = 30.0
        page.column_dimensions["H"].width = 20.0
        page.column_dimensions["I"].width = 20.0

        return wb, page, len(header_main)

    def write_render_header(self, wb, page, number_row):
        cells = "A{}:I{}".format(number_row, number_row)
        cellA = "A{}".format(number_row)
        page.merge_cells(cells)
        header_render = page[cellA]
        header_render.value = 'RENDER'
        header_render.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        header_render.font = openpyxl.styles.Font(size=12, bold=True)
        header_render.fill = openpyxl.styles.PatternFill('solid', fgColor="e9a5f2")
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        header_render.border = openpyxl.styles.Border(left=thin, right=thin)

        wb.save(self.path_to_save_report)

    def create_excel_file(self):
        wb = xlsxwriter.Workbook(self.path_to_save_report)
        wb.add_worksheet('Report')
        wb.close()

    def take_info(self):
        system = dict()
        video = dict()
        print('Host: {}'.format(self.host_at_moment))
        try:
            print('Obtaining information about the system..')
            #Take system 
            out = self.get_output_ssh('inxi -c0')
            if out != -1:
                system = self.parse_out_inxi(out)
            
            #  Take video
            out = self.get_output_ssh('nvidia-smi')
            if out != -1:
                video = self.parse_out_nvidia(out)
            print('OK')
            return [system, video]

        except (paramiko.ssh_exception.NoValidConnectionsError, socket.timeout) as r:
            print('SSH error on host - {}'.format(self.host_at_moment), "=> {}".format(r))
            return ["SSH ERROR", 0]
        
        except TypeError:
            return [0,0]
        
        except (KeyError, IndexError):
            return [system, 0]

        except Exception as r:
            print('Some error on host - {}'.format(self.host_at_moment), "=> {}".format(r))
            return ["ERROR", 0]

    def start_grabber(self, get_render = False):
        print(' -- IF YOU DO NOT SEE ERROR MESSAGES, THEN EVERYTHING WORKS, JUST WAIT ! --')
        self.create_excel_file()
        xls = self.write_main_and_fx_header()
        is_write_render_header = False
        for_merge_render_header = 3
        print(' -- HOST LIST --\n')
        print(self.hosts_list)
        print('\n -- START --\n')
        for i in range(len(self.hosts_list)):

            if 'render' in self.hosts_list[i] and get_render == False:
                continue
            elif 'render' in self.hosts_list[i] and get_render == True and is_write_render_header == False:
                self.write_render_header(xls[0], xls[1], for_merge_render_header)
                is_write_render_header = True

            self.host_at_moment = self.hosts_list[i]

            info = self.take_info()
            if 0 in info:
                self.write_data_excel(xls[0], xls[1], info[0], info[1], xls[2])
            else:
                self.write_data_excel(xls[0], xls[1], info[0], info[1])
            
            if self.progress_signal is not None:
                self.progress_signal.emit(i+1)
            
            for_merge_render_header += 1

        print('Complite!\nCheck {}'.format(self.path_to_save_report))

# a = Grab(None, ['render01'], "/home/vlad/vladcode/qt/plugins/grabber/report.xlsx")
# a.start_grabber(True)